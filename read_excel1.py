from openpyxl import load_workbook
# from openpyxl.utils import get_column_interval
import pandas as pd
import re

pd.set_option('display.max_columns', None)
pd.set_option('display.expand_frame_repr', False)
pd.set_option('display.max_rows', None)

def load_workbook_range(range_string, ws):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    # return pd.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))
    return pd.DataFrame(data_rows)

# wb = load_workbook(filename=r'E:\Nikhil\automation\Invoicing_automation\Leave_Tracker_Marketing_Finance_2020.xlsx',
                   # read_only=True)
# ws = wb['Tracker']
# df=load_workbook_range('C17:JU37', ws)
# header=df.iloc[0]
# df=df.iloc[1:]
# df.columns=header

start_index = 18
end_index = 37

def read_leave_tracker(start_index,end_index):
    # print(start_index-1, end_index)
    df=pd.read_excel(r'E:\Nikhil\automation\Invoicing_automation\Leave_Tracker_Marketing_Finance_2020.xlsx',
    sheet_name='Tracker',skiprows=start_index-2, nrows=(end_index-start_index+1))
    # df.columns = df.columns.str.strip()
    df=df.set_index('RACF ID')
    df=df.loc[:,'Resource Names ':]
    df=df.reset_index()
    del df['Resource Names ']
    df=df.set_index('RACF ID')
    df=df.T
    df.index = pd.to_datetime(df.index)
    return df

print(read_leave_tracker(start_index,end_index))
# print(df)
