#!/usr/bin/env python
import calendar
import datetime
import pandas as pd
import openpyxl as xl
from openpyxl import load_workbook
from collections import defaultdict
from openpyxl import workbook

#Method to copy cells which required starting row, column, ending row, column and sheet name as input
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    for i in range(startRow,endRow + 1,1):
        rowSelected = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
        rangeSelected.append(rowSelected)
    return rangeSelected

#Method to paste cells which required starting row, column, ending row, column and sheet name and copied data as input
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving,copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

#Method to traverse through the cells for calculation of required fields.
def traverseLeaveTracker(min_row, max_row, max_col, min_col,cell_type,col_num):
    for row in tracker.iter_rows(min_row=min_row,max_row=max_row,max_col=max_col,min_col=min_col):
        s=0
        for cell in row:
               if  cell.value == cell_type and cell_type == "W-B":
                   s = s + 1
                   invoice_sheet.cell(row=start_index_op, column=col_num).value = -(s *  hours_billable)
               elif cell.value == cell_type:
                   s = s + 1
                   invoice_sheet.cell(row=start_index_op, column=col_num).value = s * hours_billable
               else:
                   invoice_sheet.cell(row=start_index_op, column=col_num).value = s * hours_billable

#Method to calculate leave hours.
def leaveHoursCount(min_row,max_row,max_col,min_col):
    for row in tracker.iter_rows(min_row=min_row,max_row=max_row,max_col=max_col,min_col=min_col):
        s=0
        for cell in row:
               if  cell.value == 'LEA-P' or cell.value == 'LEA-U' or cell.value == 'LEA-C':
                   s = s + 1
        leaveHours = s * hours_billable
        return leaveHours


#Paths to excel sheets
wb1= xl.load_workbook(r'E:\Nikhil\automation\Invoicing_automation\Leave_Tracker_Marketing_Finance_2020.xlsx',data_only=2)
#pathLeaveTracker = 'C:\\Users\\ravit\\Desktop\\Invoicing\\Leave_Tracker_Marketing_Finance_2020.xlsx'
excecution_sheet = wb1["Invoicing_Excecution"]
temp = excecution_sheet.cell(row=6,column=2).internal_value
#pathFinalInvoicingSheet ='C:\\Users\\ravit\\Desktop\\Invoicing\\Invoicing_final_template.xlsx'
pathFinalInvoicingSheet = excecution_sheet.cell(row=5,column=2).internal_value
#pathFinalInvoicingSheet = temp
#Read output invoice sheet for updating cell values
template = xl.load_workbook(temp)
invoice_sheet = template["Final Sheet"]
tracker = wb1["Tracker"]


#print(Final_Template)
def getIndex(workbook,cell_value,type):
  for row in workbook.iter_cols():
    for cell in row:
      if cell.value == cell_value:
        if type == 'row':
          start_index= cell.row
        elif type == 'column':
          start_index= cell.col_idx
  return start_index


def getStartDate(counter):
  start_date = calendar_dict[list[counter]]['actual_start']
  return start_date

def getEndDate(counter):
  end_date = calendar_dict[list[counter]]['actual_end']
  return end_date

#Get count of total number of employees.
start_index = getIndex(tracker,"Resource Names ",'row') + 1
start_index_op = getIndex(invoice_sheet,"RACF ID",'row') + 1
start_ind=start_index


#Calculate End Index from leave tracker.
i = start_index
while True:
    if tracker.cell(row = i, column=4).value == None:
        end_index = i - 1
        break
    else:
        i = i + 1

#Get values needed for calculations.
rate = 39.89
hours_billable = 8
no_of_employees = end_index - start_index + 1
count = no_of_employees

calendar_dict={} #[type: ([billing_period_start:end], [actual_period_start:end], future,billable flag)]
def get_ranges(year,month):
#get main range - last day and future days
  s_day = max(max(week[calendar.FRIDAY]
                    for week in calendar.monthcalendar(year, month)),
                 max(week[calendar.SATURDAY]
                    for week in calendar.monthcalendar(year, month)),
                 max(week[calendar.SUNDAY]
                    for week in calendar.monthcalendar(year, month)))
  last_day=datetime.datetime.strptime('{:02d}-{:02d}-{:4d}'.format(s_day,month,year), '%d-%m-%Y')
  fut=calendar.monthrange(last_day.year,(last_day.month))[1]-last_day.day
#get main range - end day
  if calendar.weekday(year, month, 1) in [5,6,0]:
    first_day=datetime.datetime.strptime('{:02d}-{:02d}-{:4d}'.format(1, month, year), '%d-%m-%Y')
  else:
    if month==1:
      month=12
      year-=1
    else:
      month-=1
    monday = max(week[calendar.MONDAY]
                    for week in calendar.monthcalendar(year, month))
    first_day=datetime.datetime.strptime('{:02d}-{:02d}-{:4d}'.format(monday,month,year), '%d-%m-%Y')
  inner_dict={}
  inner_dict['bill_start']=datetime.datetime.strptime('{:02d}-{:02d}-{:4d}'.format(1,last_day.month,last_day.year), '%d-%m-%Y')
  inner_dict['bill_end']=datetime.datetime.strptime('{:02d}-{:02d}-{:4d}'.format(calendar.monthrange(last_day.year,last_day.month)[1],last_day.month,last_day.year), '%d-%m-%Y')
  inner_dict['actual_start']=first_day
  inner_dict['actual_end']=last_day
  inner_dict['future']=fut
  inner_dict['bill_flag']='N'
  calendar_dict['main_range']=inner_dict
  inner_dict={}
#get sub ranges
  if first_day.month == last_day.month:
    inner_dict['bill_start']=datetime.datetime.strptime('{:02d}-{:02d}-{:4d}'.format(1,first_day.month,first_day.year), '%d-%m-%Y')
    inner_dict['bill_end']=datetime.datetime.strptime('{:02d}-{:02d}-{:4d}'.format(calendar.monthrange(first_day.year,first_day.month)[1],first_day.month,first_day.year), '%d-%m-%Y')
    inner_dict['actual_start']=first_day
    inner_dict['actual_end']=last_day
    inner_dict['future']=fut
    inner_dict['bill_flag']='Y'
    calendar_dict['invoice_range']=inner_dict
  else:
    inner_dict['bill_start']=datetime.datetime.strptime('{:02d}-{:02d}-{:4d}'.format(1,first_day.month,first_day.year), '%d-%m-%Y')
    inner_dict['bill_end']=datetime.datetime.strptime('{:02d}-{:02d}-{:4d}'.format(calendar.monthrange(first_day.year,first_day.month)[1],first_day.month,first_day.year), '%d-%m-%Y')
    inner_dict['actual_start']=first_day
    inner_dict['actual_end']=datetime.datetime.strptime('{:02d}-{:02d}-{:4d}'.format(calendar.monthrange(first_day.year,first_day.month)[1],first_day.month,first_day.year), '%d-%m-%Y')
    inner_dict['future']=0
    inner_dict['bill_flag']='Y'
    calendar_dict['overlap_range']=inner_dict
    inner_dict={}
    inner_dict['bill_start']=datetime.datetime.strptime('{:02d}-{:02d}-{:4d}'.format(1,last_day.month,last_day.year), '%d-%m-%Y')
    inner_dict['bill_end']=datetime.datetime.strptime('{:02d}-{:02d}-{:4d}'.format(calendar.monthrange(last_day.year,last_day.month)[1],last_day.month,last_day.year), '%d-%m-%Y')
    inner_dict['actual_start']=datetime.datetime.strptime('{:02d}-{:02d}-{:4d}'.format(1,last_day.month,last_day.year), '%d-%m-%Y')
    inner_dict['actual_end']=last_day
    inner_dict['future']=fut
    inner_dict['bill_flag']='Y'
    calendar_dict['invoice_range']=inner_dict
  inner_dict={}
  inner_dict['bill_start']=datetime.datetime.strptime('{:02d}-{:02d}-{:4d}'.format(1,last_day.month+1,last_day.year), '%d-%m-%Y')
  inner_dict['bill_end']=datetime.datetime.strptime('{:02d}-{:02d}-{:4d}'.format(calendar.monthrange(last_day.year,last_day.month+1)[1],last_day.month+1,last_day.year), '%d-%m-%Y')
  inner_dict['actual_start']=datetime.datetime.strptime('{:02d}-{:02d}-{:4d}'.format(1,last_day.month+1,last_day.year), '%d-%m-%Y')
  inner_dict['actual_end']=datetime.datetime.strptime('{:02d}-{:02d}-{:4d}'.format(calendar.monthrange(last_day.year,last_day.month+1)[1],last_day.month+1,last_day.year), '%d-%m-%Y')
  inner_dict['future']=str(calendar.monthrange(last_day.year,(last_day.month+1))[1])
  inner_dict['bill_flag']='N'
  calendar_dict['future_range']=inner_dict
  return calendar_dict

calendar_dict=get_ranges(excecution_sheet.cell(row=1,column=2).value,excecution_sheet.cell(row=2,column=2).value)
#calendar_dict=get_ranges(excecution_sheet.cell(row=1,column=2).value,5)
print(str(excecution_sheet.cell(row=1,column=2).value) + '-' + str(excecution_sheet.cell(row=2,column=2).value))


legends={}
for i in range (1,start_ind-1):
  if tracker.cell(row=i,column=10).value!=None and tracker.cell(row=i,column=10).value!='HOL':
    legends[tracker.cell(row=i,column=10).value]=tracker.cell(row=i,column=9).value
  if tracker.cell(row=i,column=8).value!=None and tracker.cell(row=i,column=8).value!='HOL':
    legends[tracker.cell(row=i,column=8).value]=tracker.cell(row=i,column=7).value
  if tracker.cell(row=i,column=12).value!=None and tracker.cell(row=i,column=12).value!='HOL':
    legends[tracker.cell(row=i,column=12).value]=tracker.cell(row=i,column=11).value

#Writing comment for a racfid for given date range
def get_comments(start_ind,start_date,end_date,racf_row_index):
  month_start_column=getIndex(tracker,start_date,'column')
  month_end_column=getIndex(tracker,end_date,'column')
  #Comment creation
  off_dict={}
  refined_off_dict={}
  comment=''
  for j in range (month_start_column,month_end_column+1):
    if tracker.cell(row=racf_row_index,column=j).value !=None and tracker.cell(row=racf_row_index,column=j).value in legends.keys():
      #print(racf_row_index)
      off_dict[tracker.cell(row=start_ind-1,column=j).value.strftime("%d-%b")]=tracker.cell(row=racf_row_index,column=j).value
      #print(tracker.cell(row=racf_row_index,column=j).value)
  for key, value in off_dict.items():
    if value in refined_off_dict:
      refined_off_dict[value].append(key)
    else:
      refined_off_dict[value]=[key]
  #print(refined_off_dict)
  for key,value in refined_off_dict.items():
    dates=''
    for v in value:
      if v is None:
        comment=''
      else:
        dates=dates+v+' '
    comment=comment+legends[key]+' : '+dates+"\n"
  return comment

#Method to calculate number of working_days
def no_of_days (start_index,start_ind,start_date,end_date):
  bp=0
  month_start=getIndex(tracker,start_date,'column')
  month_end=getIndex(tracker,end_date,'column')
  for i in range (month_start,month_end+1):
    if tracker.cell(row=start_ind-2,column=i).value not in ('Sun','Sat'):
      if tracker.cell(row=start_index,column=i).value !='HOL':
        bp+=1
  return bp


list = []
for key in calendar_dict.keys():
  list.append(key)

#Method to Calculate required fields
def Calculate(start_index,start_date,end_date):
    selectedRange = copyRange(4,start_index,10,start_index,tracker)
    pasteRange(1,start_index_op,7,start_index_op,invoice_sheet,selectedRange)
    traverseLeaveTracker(start_index,start_index,end_date_column_index,start_date_column_index,"HOL",15)
    traverseLeaveTracker(start_index,start_index,end_date_column_index,start_date_column_index,"TRG",18)
    traverseLeaveTracker(start_index,start_index,end_date_column_index,start_date_column_index,"TRV",19)
    traverseLeaveTracker(start_index,start_index,end_date_column_index,start_date_column_index,"W-B",21)
    #invoice_sheet.cell(row=start_index_op, column=14).value = invoice_sheet.cell(row=start_index_op, column=12).value - (invoice_sheet.cell(row=start_index_op, column=13).value + invoice_sheet.cell(row=start_index_op, column=15).value + invoice_sheet.cell(row=start1, column=16).value + temp_sheet.cell(row=start1, column=17).value + temp_sheet.cell(row=start1, column=18).value + temp_sheet.cell(row=start1, column=19).value + temp_sheet.cell(row=start1, column=20).value)
    #invoice_sheet.cell(row=start_index_op, column=23).value = '$'+str(rate * invoice_sheet.cell(row=start_index_op, column=13).value)
    invoice_sheet.cell(row=start_index_op, column=8).value = calendar_dict[list[counter]]['actual_start'].strftime("%d-%b") + ' to ' + calendar_dict[list[counter]]['actual_end'].strftime("%d-%b")
    invoice_sheet.cell(row=start_index_op, column=9).value = calendar_dict[list[counter]]['bill_start'].strftime("%d-%b") + ' to ' + calendar_dict[list[counter]]['bill_end'].strftime("%d-%b")
    invoice_sheet.cell(row=start_index_op, column=16).value = leaveHoursCount(start_index,start_index,end_date_column_index,start_date_column_index)
    invoice_sheet.cell(row=start_index_op, column=17).value = calendar_dict[list[counter]]['future']
    invoice_sheet.cell(row=start_index_op, column=10).value = no_of_days(start_index,start_ind,calendar_dict[list[counter]]['bill_start'],calendar_dict[list[counter]]['bill_end'] )
    invoice_sheet.cell(row=start_index_op, column=11).value = no_of_days(start_index,start_ind,calendar_dict[list[counter]]['actual_start'],calendar_dict[list[counter]]['actual_end'] )
    invoice_sheet.cell(row=start_index_op, column=12).value = invoice_sheet.cell(row=start_index_op, column=11).value * 8
    invoice_sheet.cell(row=start_index_op, column=22).value = get_comments(start_ind,calendar_dict[list[counter]]['actual_start'],calendar_dict[list[counter]]['actual_end'],start_index)

invoice_sheet.cell(row=1, column=1).value = str(excecution_sheet.cell(row=1,column=2).value) + '-' + calendar.month_name[excecution_sheet.cell(row=2,column=2).value]
#Main Loop
while (count > 0):
  counter = 1
  #num_of_months is the number of iteration inner loop goes through
  num_of_months = len(calendar_dict) - 1
  while (num_of_months > 0):
    #Gives the actual start date from dictionary
    start_date = getStartDate(counter)
    #Gives the actual end date from dictionary
    end_date = getEndDate(counter)
    #Gives the column index of actual start date from Leave Tracker
    start_date_column_index = getIndex(tracker,start_date,'column')
    #Gives the column index of actual end date from Leave Tracker
    end_date_column_index = getIndex(tracker,end_date,'column')
    #print(end_date_column_index)

    Calculate(start_index,start_date,end_date)
    start_index_op = start_index_op + 1
    counter = counter + 1
    num_of_months = num_of_months - 1
  start_index = start_index + 1
  count = count - 1

#Save changes updated in the final sheet.
template.save(pathFinalInvoicingSheet)
print("hello")
