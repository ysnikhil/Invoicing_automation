import pandas as pd
from openpyxl import load_workbook
import re

pd.set_option('display.max_columns',None)
pd.set_option('display.expand_frame_repr', False)
pd.set_option('display.max_rows', None)

vmsdump_df = pd.read_excel(r"E:\Nikhil\automation\Invoicing_automation\vms_dump.xlsx", header=0, sheet_name='Sheet2')
vmsdump_df = vmsdump_df.set_index('RACF ID')    #Setting index to Racf id to make calculations easier
vmsdump_df = vmsdump_df[['WeekEnding', 'Reg Hours', 'OT Hours']]  #Dropped the unused columns
working_hrs_per_day = 8
print(vmsdump_df)

# Function to read the excel in openpyxl and then transfer the data into a Pandas DataFrame
def load_workbook_range(range_string, ws):
    col_start, col_end = re.findall("[A-Z]+", range_string)

    data_rows = []
    for row in ws[range_string]:
        data_rows.append([cell.value for cell in row])

    # return pd.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))
    return pd.DataFrame(data_rows)

leave_tracker_full = load_workbook(filename=r'E:\Nikhil\automation\invoicing\Leave_Tracker_Marketing_Finance_2020.xlsx',
                   read_only=True)
active_sheet = leave_tracker_full['Tracker']
leave_tracker_df=load_workbook_range('C17:JU37', active_sheet)  #Pass the range of the data to be read from openpyxl to Pandas

# Make the First row of the DataFrame as the header. This is not directly achievable with the current function,
# as we need a way to read the first row in openpyxl and pass it as columns. To be optimized later.
header=leave_tracker_df.iloc[0]
leave_tracker_df=leave_tracker_df.iloc[1:]
leave_tracker_df.columns=header
# print(leave_tracker_df)

# ===================================================================================================================
# Reads the user Input and then generates the calender days expected for the current month Invoicing.
# This is needed as some teammates might take leave or submit 0 hours in the VMS and we don't get VMS data for those
# To handle those cases we are first creating a perfect calender and then fill in details from VMS Dump.
# In case any week is not available in the VMS Dump sheet, then that week would become 0.
# ===================================================================================================================

def create_default_calender(racf_id,start_date,end_date):
    date_index=pd.date_range(start=start_date, end=end_date,freq='W')
    vms_generated_calndr_df=pd.DataFrame(date_index, columns=['WeekEnding'])
    vms_generated_calndr_df['RACF ID'] = racf_id
    vms_generated_calndr_df = vms_generated_calndr_df.set_index('RACF ID')
    # print (vms_generated_calndr_df)

    return vms_generated_calndr_df




# Reads the VMS Dump and create a pandas DataFrame with needed columns
def generate_vms_sheet(racf_id,vms_generated_calndr_df):    #(leave_tracker_index,racf_id,start_date,end_date):
    vmsdump_user_df = pd.merge(vms_generated_calndr_df,vmsdump_df,how='left',on=['WeekEnding'])
    vmsdump_user_df['vms_WeekStarting'] = vmsdump_user_df['WeekEnding'] + pd.offsets.Day(-6)
    vmsdump_user_df[['Reg Hours','OT Hours']] = vmsdump_user_df[['Reg Hours','OT Hours']].fillna(0) #Replace NaN with 0
    vmsdump_user_df['vms_pending_hours'] = vmsdump_user_df['Reg Hours'] + vmsdump_user_df['OT Hours']  #Created a new column for keeping VMS hours counter
    vmsdump_user_df['vms_working_days'] = (vmsdump_user_df['Reg Hours'] + vmsdump_user_df['OT Hours']) / working_hrs_per_day

    # Below code resample the VMS Weekly data into Daily data.
    # As the resample method doesn't expand the last entry till the end, so we have to add another duplicate last row for the same.
    vmsdump_user_df = vmsdump_user_df.append(vmsdump_user_df.iloc[-1])  #appends the last row again
    vmsdump_user_df.iloc[-1, vmsdump_user_df.columns.get_loc('vms_WeekStarting')] = vmsdump_user_df.iloc[-1, vmsdump_user_df.columns.get_loc('WeekEnding')]
    # vmsdump_user_df = vmsdump_user_df.reset_index().set_index('vms_WeekStarting').resample('D').ffill().reset_index().set_index('RACF ID')

    print(vmsdump_user_df)

# vms_generated_calndr_df=create_default_calender('UGAM211','2020-03-30','2020-04-26')
# generate_vms_sheet('UINY1',vms_generated_calndr_df)
